import csv
import openpyxl
from bose.utils import write_json, read_json

class Output:

    def read_json(filename):
        if not filename.startswith("output/"):
            filename = "output/" +  filename

        if not filename.endswith(".json"):
            filename = filename + ".json"

        return read_json(filename)

    def write_json(data, filename):
        if len(data) == 0:
            print("Data is empty.")
            return

        if not filename.startswith("output/"):
            filename = "output/" +  filename

        if not filename.endswith(".json"):
            filename = filename + ".json"

        write_json(data, filename)
        print(f"View written JSON file at {filename}")        

    def write_csv(data, filename):
        """
        Save a list of dictionaries as a CSV file.

        Args:
            data: a list of dictionaries
            filename: the name of the CSV file to save
        """
        if len(data) == 0:
            print("Data is empty.")
            return


        if not filename.startswith("output/"):
            filename = "output/" +  filename

        if not filename.endswith(".csv"):
            filename = filename + ".csv"

        with open(filename, 'w', newline='') as csvfile:
            fieldnames = data[0].keys()  # get the fieldnames from the first dictionary
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()  # write the header row
            writer.writerows(data)  # write each row of data
        print(f"View written CSV file at {filename}")        
    def write_xlsx(data, filename):
        """
        Saves a list of dictionaries as an Excel file with the given filename.
        Each dictionary in the list corresponds to a row in the Excel file.
        The keys of each dictionary correspond to column names in the Excel file.
        """
        if len(data) == 0:
            print("Data is empty.")
            return

        if not filename.startswith("output/"):
            filename = "output/" +  filename

        if not filename.endswith(".xlsx"):
            filename = filename + ".xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active

        # write headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num).value = header

        # write data
        if data:
            for row_num, data in enumerate(data, start=2):
                for col_num, key in enumerate(headers, start=1):
                    ws.cell(row=row_num, column=col_num).value = data.get(key)

        # save file
        wb.save(filename)
        print(f"View written Excel file at {filename}")        

    def read_pending():
        return Output.read_json("pending.json")

    def write_pending(data):
        return Output.write_json(data, "pending.json")

    def read_finished():
        return Output.read_json("finished.json")

    def write_finished(data):
        return Output.write_json(data, 'finished.json')
