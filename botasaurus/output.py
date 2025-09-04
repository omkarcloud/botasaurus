from json import dumps
from .decorators_utils import create_directory_if_not_exists
from .utils import (
    read_file as _read_file,
    relative_path,
    write_json as _write_json,
    read_json as _read_json,
    write_html as _write_html,
    write_file as _write_file,
)
from .beep_utils import prompt
import os

def is_slash_not_in_filename(filename):
    return "/" not in filename and "\\" not in filename


_output_directory = "output/"
def get_output_directory():
    """
    Get the current output directory.
    """
    return _output_directory

def set_output_directory(directory):
    """
    Set a new output directory.
    :param directory: The new directory to set as the output directory.
    """
    global _output_directory
    _output_directory = directory

output_check_done = False
def create_output_directory_if_not_exists():
    global output_check_done
    if not output_check_done:
        output_check_done = True
        create_directory_if_not_exists(get_output_directory())
        
def append_output_if_needed(filename):
    create_output_directory_if_not_exists()
    filename = str(filename).strip()
    if is_slash_not_in_filename(filename):
        return get_output_path(filename)
    return filename

def get_output_path(filename):
    return os.path.join(_output_directory, filename)


def fix_json_filename(filename):
    filename = append_output_if_needed(filename)

    if not filename.endswith(".json"):
        filename = filename + ".json"
    return filename


def read_json(filename):
    filename = fix_json_filename(filename)

    return _read_json(filename)



def write_temp_json(data, log=True):
    filename = "temp"

    try:

        filename = append_output_if_needed(filename)

        if not filename.endswith(".json"):
            filename = filename + ".json"

        _write_json(data, filename)

        if log:
            print(f"View written JSON file at {filename}")
    except PermissionError:
        prompt(
            f"{filename} is currently open in another application. Please close the the Application and press 'Enter' to save."
        )
        return write_json(data, filename, log)

    return filename
def read_temp_json():
    filename = fix_json_filename("temp")
    return _read_json(filename)


def file_exists(filename):
    filename = append_output_if_needed(filename)

    try:
        with open(filename, "r"):
            return True
    except FileNotFoundError:
        return False


def write_json(data, filename, log=True):
    # if type(data) is list and len(data) == 0:
    #     # if log:
    #     print("No JSON File written as data list is empty.")
    #     return

    try:

        filename = append_output_if_needed(filename)

        if not filename.endswith(".json"):
            filename = filename + ".json"

        _write_json(data, filename)

        if log:
            print(f"View written JSON file at {filename}")
        
    except PermissionError:
        prompt(
            f"{filename} is currently open in another application. Please close the the Application and press 'Enter' to save."
        )
        return write_json(data, filename, log)
    return filename
def write_temp_csv(data, log=True):
    return write_csv(data, "temp.csv", log)


def read_temp_csv():
    return read_csv("temp.csv")


def fix_csv_filename(filename):
    filename = append_output_if_needed(filename)

    if not filename.endswith(".csv"):
        filename = filename + ".csv"
    return filename


def read_csv(filename):
    """
    Reads a CSV file and returns a list of dictionaries.

    :param filepath: str, the path to the CSV file
    :return: list of dictionaries
    """
    import csv

    filename = fix_csv_filename(filename)

    data = []
    with open(filename, mode="r") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            data.append(row)
    return data

def get_fieldnames(data_list):
    fieldnames_dict = {}  # Initialize an empty dictionary
    for item in data_list:
        for key in item.keys():
            if key not in fieldnames_dict:
                fieldnames_dict[key] = (
                    None  # Set the value to None, we only care about the keys
                )
    return list(fieldnames_dict.keys())  # Convert the dictionary keys to a list

def convert_nested_to_json(input_list):
    """
    Iterates through a list of dictionaries and converts any nested dictionaries or lists
    within those dictionaries into JSON-formatted strings.

    :param input_list: The list of dictionaries to process.
    :return: A new list with dictionaries having nested dictionaries/lists converted to JSON strings.
    """
    output_list = []

    for item in input_list:
        processed_dict = {}
        for key, value in item.items():
            if isinstance(value, (dict, list, tuple, set)):
                # Convert the value to a JSON-formatted string if it's a dict or list
                processed_dict[key] = dumps(value)
            else:
                # Keep the value as is if it's not a dict or list
                processed_dict[key] = value
        output_list.append(processed_dict)

    return output_list

def cap(string):
    """
    Truncate the string if it exceeds 32767 characters and append '...'.
    
    :param string: The string to truncate.
    :return: The truncated string.
    """
    if len(string) > 32767:
        return string[:32764] + '...'
    return string

def convert_nested_to_json_for_excel(input_list):
    """
    Iterates through a list of dictionaries and converts any nested dictionaries or lists
    within those dictionaries into JSON-formatted strings. Also truncates long strings.

    :param input_list: The list of dictionaries to process.
    :return: A new list with dictionaries having nested dictionaries/lists converted to JSON strings.
    """
    output_list = []

    for item in input_list:
        processed_dict = {}
        for key, value in item.items():
            if isinstance(value, (dict, list, tuple, set)):
                # Convert the value to a JSON-formatted string if it's a dict or list
                processed_dict[key] = cap(dumps(value))
            elif isinstance(value, str):
                # Truncate the string if it's too long
                processed_dict[key] = cap(value)
            else:
                # Keep the value as is if it's not a dict or list
                processed_dict[key] = value
        output_list.append(processed_dict)

    return output_list

def remove_non_dicts(data):
    return [x for x in data if isinstance(x, dict)]

def write_csv(data, filename, log=True):
    """
    Save a list of dictionaries as a CSV file.

    Args:
        data: a list of dictionaries
        filename: the name of the CSV file to save
    """
    import csv

    data = clean_data(data)
    data = convert_nested_to_json(data)

    filename_new = append_output_if_needed(filename)


    if not filename_new.endswith(".csv"):
        filename_new = filename_new + ".csv"

    try:
        with open(filename_new, "w", newline="", encoding="utf-8") as csvfile:
            fieldnames = get_fields(data)
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()  # write the header row
            for i in data:
                writer.writerow(i)  # write each row of data
        if log:
            print(f"View written CSV file at {filename_new}")

    except PermissionError:
        prompt(
            f"{filename_new} is currently open in another application (e.g., Excel). Please close the the Application and press 'Enter' to save."
        )
        return write_csv(data, filename, log)
    return filename_new

def save_image(url, filename=None):
    import requests

    if filename is None:
        filename = url.split("/")[-1]
    response = requests.get(url)
    if response.status_code == 200:

        # Save the image in the output directory
        path = append_output_if_needed(filename)
        with open(relative_path(path), "wb") as f:
            f.write(response.content)
    else:
        print("Failed to download the image.")
        return path


def write_html(data, filename, log=True):
    filename = append_output_if_needed(filename)

    if not filename.endswith(".html"):
        filename = filename + ".html"

    _write_html(data, filename)
    if log:
        print(f"View written HTML file at {filename}")
    return filename

def read_html(filename):
    filename = append_output_if_needed(filename)

    if not filename.endswith(".html"):
        filename = filename + ".html"

    return _read_file(filename)


def write_temp_html(data, log=True):
   return write_html(data, "temp.html", log)


def read_temp_html():
    return read_html("temp.html")

def write_file(data, filename, log=True):

    filename = append_output_if_needed(filename)

    _write_file(data, filename)
    if log:
        print(f"View written file at {filename}")

    return filename
def read_file(filename):
    filename = append_output_if_needed(filename)

    return _read_file(filename)

def write_temp_file(data, log=True):
   return write_file(data, "temp.txt", log)

def read_temp_file():
    return read_file("temp.txt")
def normalize_data(data):
    if data is None:
        return []
    elif isinstance(data, (list, set, tuple)):
        normalized_list = []
        for item in data:
            if item is None:
                continue  # Skip None items
            elif not isinstance(item, dict):
                item = {"data": item}  # Wrap non-dict items
            normalized_list.append(item)
        return normalized_list

    elif isinstance(data, dict):
        return [data]

    else:
        return [{"data": data}]
    
def normalize_dicts_by_fieldnames(data):
    fieldnames = get_fieldnames(data)
    filtered_data = []

    for item in data:
        filtered_item = {key: item.get(key, None) for key in fieldnames}
        filtered_data.append(filtered_item)

    return filtered_data

def clean_data(data):
    # Then, filter the normalized data to ensure each dict contains the same set of fieldnames
    return normalize_dicts_by_fieldnames(normalize_data(data))

def fix_excel_filename(filename):
    filename = append_output_if_needed(filename)

    if not filename.endswith(".xlsx"):
        filename = filename + ".xlsx"
    return filename
def write_excel(data, filename, log=True, convert_strings_to_urls=True):

    data = clean_data(data)
    data = convert_nested_to_json_for_excel(data)

    try:
        filename = fix_excel_filename(filename)
        write_workbook(data, filename, convert_strings_to_urls)

        if log:
            print(f"View written Excel file at {filename}")
    except PermissionError:
        prompt(f"{filename} is currently open in another application (e.g., Excel). Please close the the Application and press 'Enter' to save.")
        return write_excel(data, filename, log, convert_strings_to_urls)
    return filename
MAX_EXCEL_LINKS = 65528
def write_workbook(data, filename,  strings_to_urls = True):
    import xlsxwriter
    if strings_to_urls:
     workbook = xlsxwriter.Workbook(filename)
    else:
     workbook = xlsxwriter.Workbook(filename, {'strings_to_urls': False})
     
    worksheet = workbook.add_worksheet()

        # Write headers
    fieldnames = get_fields(data)
    worksheet.write_row(0, 0, fieldnames)

        # Write data
    row = 1
    for item in data:
        # Prevent Warnings and Handle Excel 65K Link Limit
        if worksheet.hlink_count > MAX_EXCEL_LINKS:
            workbook.close()
            if os.path.exists(filename):
                os.remove(filename)
            return write_workbook(data, filename,strings_to_urls = False)
        values = list(item.values())
        worksheet.write_row(row, 0, values)
        row += 1

    workbook.close()
    return filename
def get_fields(data):
    if len(data) == 0:
        return []
    return list(data[0].keys())

def read_excel(filename):
    import openpyxl

    filename = fix_excel_filename(filename)

    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    data = []
    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2):
        row_data = {}
        for cell, header in zip(row, headers):
            row_data[header] = cell.value
        data.append(row_data)

    return data

def write_temp_excel(data, log=True, convert_strings_to_urls=True):
    return write_excel(data, "temp.xlsx", log, convert_strings_to_urls)

def read_temp_excel():
    return read_excel("temp.xlsx")

def zipdir(path, ziph):
    # ziph is zipfile handle
    for root, dirs, files in os.walk(path):
        for file in files:
            
            main = os.path.join(root, file)
            arch = os.path.relpath(os.path.join(root, file), os.path.join(path, '..'))
            ziph.write(main, arch)
            

def zip_files(filenames, output_filename=None, log=True):
    import zipfile

    # Ensure filenames is a list
    if not isinstance(filenames, list):
        filenames = [filenames]

    # Filter out any None or empty string filenames
    filenames = [f for f in filenames if f]

    # If no valid filenames, return early
    if not filenames:
        print("No files to zip.")
        return None
    if not output_filename: 
        if len(filenames) == 1:
            output_filename = "./" + os.path.basename(filenames[0]) + ".zip"
        else: 
            output_filename = "./data.zip"


    # Prepare the output filename
    if not output_filename.endswith(".zip"):
        output_filename += ".zip"

    if is_slash_not_in_filename(output_filename):
        output_filename = "./"+output_filename.strip()
    else: 
        output_folder = os.path.dirname(output_filename)
        os.makedirs(output_folder, exist_ok=True)

    try:
        with zipfile.ZipFile(output_filename, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
            for file in filenames:
                # file = append_output_if_needed(file)
                file_exists = os.path.exists(file)
                if file_exists:
                    if os.path.isdir(file):
                        zipdir(file, zipf)
                    else: 
                        # Get the file name without the "output/" prefix
                        arcname =  os.path.basename(file) 
                        zipf.write(file, arcname=arcname)
                else:
                    raise Exception(f"{file}' not found. Unable to add to zip archive.")

        if log:
            print(f"View written ZIP file at {output_filename}")
        return output_filename

    except PermissionError:
        prompt(f"{output_filename} is currently open in another application. Please close the application and press 'Enter' to retry.")
        return zip_files(filenames, output_filename, log)
    except zipfile.BadZipFile:
        print(f"Error: Unable to create zip file {output_filename}. It may be corrupted.")
        return None
    except Exception as e:
        print(f"Error while zipping files: {e}")
        return None
    
def _has_less_than1_item(path):
    file_count = 0
    for _ in os.scandir(path):
        file_count += 1
            
        if file_count > 1:
            return False
    
    return True

def _has_0_item(path):
    file_count = 0
    for _ in os.scandir(path):
        file_count += 1
        return False
    
    return True

def unzip_file(filename, output_folder_path=None, force=False, log=True):
    import zipfile
    import shutil

    # Ensure filename is provided
    if not filename:
        print("No file provided to unzip.")
        return None

    # Ensure the file exists
    if not os.path.exists(filename):
        print(f"File '{filename}' not found.")
        return None

    # Ensure the file has a .zip extension
    if not filename.endswith(".zip"):
        print(f"File '{filename}' is not a zip file.")
        return None
    
    # Ensure the file is a zip file
    if not zipfile.is_zipfile(filename):
        print(f"Error: '{filename}' is not a valid zip file.")
        return None
    
    try:
        # Extract the base name of the file (without the .zip extension)
        base_name = os.path.splitext(os.path.basename(filename))[0]
        if output_folder_path is None:
            output_folder_path = os.path.join(os.getcwd(), base_name)
        else: 
            output_folder_path = output_folder_path.rstrip("/")

        # Check if the output folder already exists
        if os.path.exists(output_folder_path):
            if _has_0_item(output_folder_path):
                pass
            else:
                if not force:
                    print(f"Error: Folder '{output_folder_path}' already exists.")
                    return None
        
        os.makedirs(output_folder_path, exist_ok=True)

        with zipfile.ZipFile(filename, 'r') as zipf:
            zipf.extractall(output_folder_path)


        folder_in_folder = os.path.join(output_folder_path, os.path.basename(output_folder_path))
        is_folder_in_folder = os.path.isdir(folder_in_folder)
        
        if is_folder_in_folder:
                result = _has_less_than1_item(output_folder_path)
                if result:
                    s = folder_in_folder
                    d = os.path.join(os.path.dirname(filename), "__" + base_name)
                    real_dst = shutil.move(s, d)        
                    shutil.rmtree(output_folder_path)    
                    os.rename(real_dst, base_name)   

        if log:
            if output_folder_path == os.path.join(os.getcwd(), os.path.basename(output_folder_path)):
              bn = os.path.basename(output_folder_path)
              print(f"Unzipped files can be found in the folder: ./{bn}")
            else:
              print(f"Unzipped files can be found in the folder: {output_folder_path}")
            
        return output_folder_path

    except PermissionError:
        prompt(f"{filename} is currently open in another application. Please close the application and press 'Enter' to retry.")
        return unzip_file(filename, log)
    except zipfile.BadZipFile:
        print(f"Error: Unable to unzip file {filename}. It may be corrupted.")
        return None
    except Exception as e:
        print(f"Error while unzipping file: {e}")
        return None


def get_metadata(file_name):
    """
    Determine metadata based on file extension.
    
    :param file_name: The file name or path
    :return: A dictionary with appropriate metadata
    """
    extension = os.path.splitext(file_name)[1].lower()

    extension_mapping = {
        '.csv': {'Content-Type': 'text/csv'},
        '.json': {'Content-Type': 'application/json'},
        '.zip': {'Content-Type': 'application/zip'},
        '.xlsx': {'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'},
        '.html': {'Content-Type': 'text/html'},
        '.txt': {'Content-Type': 'text/plain'},
        '.dmg': {'Content-Type': 'application/x-apple-diskimage'},
        '.exe': {'Content-Type': 'application/x-msdownload'},
        '.deb': {'Content-Type': 'application/vnd.debian.binary-package'},
        '.rpm': {'Content-Type': 'application/x-rpm'},
        '.pdf': {'Content-Type': 'application/pdf'}
    }
    
    return extension_mapping.get(extension, None)

def install(package):
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

def dynamically_import_boto3():
    try:
        import boto3
    except ImportError:
        install('boto3')

def get_aws_access_message(access_key_id: str = None, secret_access_key: str = None) -> str:
    """
    Generate message based on missing AWS credentials.

    Args:
        access_key_id (str): AWS access key ID
        secret_access_key (str): AWS secret access key

    Returns:
        str: Appropriate message about missing credentials
    """
    if not access_key_id and not secret_access_key:
        return "AWS access key ID and secret access key are missing."
    elif not access_key_id:
        return "AWS access key ID is missing."
    elif not secret_access_key:
        return "AWS secret access key is missing."

def upload_to_s3(file_name, bucket_name, access_key_id, secret_access_key, object_name=None,):
    """
    Upload a file to an S3 bucket

    :param file_name: Local file path to upload
    :param access_key_id: AWS Access Key ID
    :param secret_access_key: AWS Secret Access Key
    :param bucket_name: S3 bucket name
    """
    if not os.path.exists(file_name):
        if not os.path.exists(append_output_if_needed(file_name)):
            raise FileNotFoundError(f"{file_name} not found. Unable to upload to S3.")
        else:
            file_name = append_output_if_needed(file_name)
    if not access_key_id or not secret_access_key:
        msg = get_aws_access_message(access_key_id, secret_access_key)
        raise ValueError(msg)
    dynamically_import_boto3()
    import boto3

    s3_client = boto3.client(
        's3',
        aws_access_key_id=access_key_id,
        aws_secret_access_key=secret_access_key
    )

    # Extract the file name to use as the S3 object name
    object_name = object_name or os.path.basename(file_name)
    metadata = get_metadata(file_name)

    with open(file_name, 'rb') as file:
        if metadata:
            s3_client.upload_fileobj(
                file,
                bucket_name,
                object_name,
                ExtraArgs={'Metadata': metadata}
            )
        else:
            s3_client.upload_fileobj(
                file,
                bucket_name,
                object_name,
            )
        print(f'Successfully uploaded file')
    
    return f"https://{bucket_name}.s3.amazonaws.com/{object_name}"

def download_from_s3(file_name, object_name, bucket_name,  access_key_id, secret_access_key, region_name='us-east-1'):
    """
    Download an object from an S3 bucket

    :param file_name: Local file path to save the downloaded object (can be relative or absolute)
    :param bucket_name: S3 bucket name
    :param object_name: S3 object name to download
    :param access_key_id: AWS Access Key ID
    :param secret_access_key: AWS Secret Access Key
    :param region_name: AWS region name (default is 'us-east-1')
    """
    
    dynamically_import_boto3()
    import boto3
    from botocore.exceptions import ClientError

    # Convert file_name to an absolute path
    file_name = os.path.abspath(file_name)
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(file_name), exist_ok=True)

    # Create an S3 client
    s3_client = boto3.client(
        's3',
        aws_access_key_id=access_key_id,
        aws_secret_access_key=secret_access_key,
        region_name=region_name
    )

    try:
        # Download the file
        s3_client.download_file(bucket_name, object_name, file_name)
        print(f'Successfully downloaded object "{object_name}" to {file_name}')
        
    except ClientError as e:
        if e.response['Error']['Code'] == "404":
            print("The object does not exist.")
        else:
            print(f"An error occurred: {e}")
        return False
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return False

    return True